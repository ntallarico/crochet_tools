"""
Copyright (c) 2025 Nicholas Tallarico

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

"""

#################################
#       Crochet Tools            
#            v1.0                
#################################

# ---------- Import Libraries ----------

import sys
from io import TextIOWrapper
from os import path, mkdir
import customtkinter as ctk
from tkinter import filedialog
from PIL import Image, ImageTk, ImageEnhance
from tkinter import messagebox
from openpyxl import styles, Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from string import ascii_uppercase

# ---------- Global Variables ----------

# Aesthetics
error_box_header = "Error"
window_title = "Crochet Pattern Image Editor"
color_base = "#217346"
window_width = 1600
window_height = 900

# Functionality
csv_output_directory = "input_output"
max_color_input = 32
min_color_input = 1
max_dimension_input = 500
min_dimension_input = 1

# ---------- Global State ----------

image_lvl0 = None # original
image_lvl1 = None
image_lvl2 = None
image_lvl3 = None





# ---------- Classes ----------

# custom stream class that will handle the redirection of standard and error output to the UI console
class TextBoxRedirector(TextIOWrapper):
    def __init__(self, textbox):
        super().__init__(sys.stdout.buffer)
        self.textbox = textbox
        self._buffer = []
        
    def write(self, string):
        self._buffer.append(string)
        self.textbox.configure(state="normal")
        self.textbox.insert("end", string)
        self.textbox.see("end")
        self.textbox.update()  # force the textbox to update immediately instead of writing things to a buffer
        
        # keep only last 1000 lines
        line_count = int(self.textbox.index('end-1c').split('.')[0])
        if line_count > 1000:
            self.textbox.delete("1.0", "2.0")
        
        self.textbox.configure(state="disabled")
        
    def flush(self):
        pass


# ---------- Functions ----------

def resize_for_display(image, max_size=500):
    if image == None: return
    # Resize image maintaining aspect ratio to fit within max_size x max_size
    ratio = min(max_size/image.width, max_size/image.height)
    new_size = (int(image.width * ratio), int(image.height * ratio))
    return image.resize(new_size, Image.Resampling.NEAREST)

def select_file():
    global image_lvl0, image_lvl1, image_lvl2, image_lvl3
    filepath = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png *.jpeg")])
    if filepath:
        image_lvl0 = Image.open(filepath).convert("RGB")
        update_image_display(image_lvl0, image_lvl0_image_label)
        update_all_levels()
        print(f"Image file loaded:\n'{filepath}'")

def reset_sliders():
    if image_lvl0 == None: return
    brightness_slider.set(slider_defaults["brightness"])
    contrast_slider.set(slider_defaults["contrast"])
    saturation_slider.set(slider_defaults["saturation"])
    #apply_color_sliders()
    update_all_levels()


# updates display for a given image. does not update it in memory
def update_image_display(image, image_label):
    if image == None: return

    display_img = resize_for_display(image, max_size=500)
    img_tk_new = ImageTk.PhotoImage(display_img)
    image_label.configure(image=img_tk_new)
    image_label.image = img_tk_new

def process_lvl0_to_lvl1():
    global image_lvl0, image_lvl1, image_lvl1_image_label
    new_image = image_lvl0.copy()
    # apply color sliders
    new_image = apply_color_sliders(new_image)
    # update image_lvl1 in memory
    image_lvl1 = new_image
    # update image_lvl1 on the UI display
    update_image_display(image_lvl1, image_lvl1_image_label)

def process_lvl1_to_lvl2():
    global image_lvl1, image_lvl2, image_lvl2_image_label

    if not dimensions_valid(width_entry.get(), height_entry.get()):
        return None, None
    if not num_colors_valid(colors_entry.get()):
        return None, None
    
    try:
        width = int(width_entry.get())
        height = int(height_entry.get())
        num_colors = int(colors_entry.get())
    except ValueError:
        print("Invalid width/height/colors")
        return
    
    new_image = image_lvl1.copy()
    # pixelate image (resize)
    new_image = pixelate_image(new_image, width, height)
    # quantize (reduce color palette)
    new_image = quantize_image(new_image, num_colors)
    # update image_lvl2 in memory
    image_lvl2 = new_image
    # update image_lvl2 on the UI display
    update_image_display(image_lvl2, image_lvl2_image_label)

# currently unused
def process_lvl2_to_lvl3():
    global image_lvl2, image_lvl3
    # curently does nothing
    new_image = image_lvl2.copy()
    # update image_lvl3 in memory
    image_lvl3 = new_image
    # update image_lvl3 on the UI display
    #update_image_display(image_lvl3, image_lvl3_image_label)

def update_all_levels():
    process_lvl0_to_lvl1()
    process_lvl1_to_lvl2()
    process_lvl2_to_lvl3()

def apply_color_sliders(image):
    if image == None: return

    brightness = brightness_slider.get()
    contrast = contrast_slider.get()
    saturation = saturation_slider.get()

    img = image.copy()
    img = ImageEnhance.Brightness(img).enhance(brightness)
    img = ImageEnhance.Contrast(img).enhance(contrast)
    img = ImageEnhance.Color(img).enhance(saturation)

    return img

def pixelate_image(image, width, height):
    if image == None: return
    image_pixelated = image.resize((width, height))
    return image_pixelated

def quantize_image(image, num_colors):
    if image == None: return
    if num_colors <= 32:
        quantized_image = image.convert("P", palette=Image.ADAPTIVE, colors=num_colors, dither=0).convert("RGB")
    return quantized_image

def dimensions_valid(width, height):
	## Check if height and width are numbers
	if not width.isnumeric():
		print("Error: Width contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Width contains non-numeric characters.")
		return False
	if not height.isnumeric():
		print("Error: Height contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Height contains non-numeric characters.")
		return False
	## Check if height in width are within the desired range
	width = int(width)
	height = int(height)
	if width < min_dimension_input or width > max_dimension_input:
		print("Error: Width '" + str(width) + "' not valid. Must be between " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		messagebox.showinfo(error_box_header, "Error: Width '" + str(width) + "' not valid. Must be between " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		return False
	if height < min_dimension_input or height > max_dimension_input:
		print("Error: Height '" + str(height) + "' not valid. Must be between " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		messagebox.showinfo(error_box_header, "Error: Height '" + str(height) + "' not valid. Must be " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		return False
	return True

def num_colors_valid(num_colors):
	## Check if number is numeric
	if not num_colors.isnumeric():
		print("Error: Number of Colors contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Number of Colors contains non-numeric characters.")
		return False
	## Check if in desired range
	num_colors = int(num_colors)
	if num_colors < min_color_input or num_colors > max_color_input:
		print("Error: Number of Colors '" + str(num_colors) + "' not valid. Must be between "  + str(min_color_input) + " and " + str(max_color_input))
		messagebox.showinfo(error_box_header, "Error: Number of Colors '" + str(num_colors) + "' not valid. Must be between "  + str(min_color_input) + " and " + str(max_color_input))
		return False
	return True

# IN: PIL image
# OUT: 2D array with each value containing a rgb tuple
# OUT: 2D array with each value containing an int representing the color map value
def get_colors(image):
	used_colors = []
	colors = []
	color_map = []

	for x in range(0, image.size[0]): # Left column to right column
		column_colors = []
		column_map = []
		for y in range(0, image.size[1]): # Top row to bottom row
			pixel_color = image.getpixel((x,y))
			if(pixel_color not in used_colors):
				used_colors.append(pixel_color)
			pixel_map = used_colors.index(pixel_color)

			column_colors.append(pixel_color)
			column_map.append(pixel_map)
		colors.append(column_colors)
		color_map.append(column_map)

	return colors, color_map

def get_font_color(cell_color):
	r = cell_color[0]
	g = cell_color[1]
	b = cell_color[2]
	luma = 0.299*r + 0.587*g + 0.114*b
	luma = luma / 255.0 # Account for rgb scale being 0-255 instead of 0-1.0
	if luma > 0.7: # Cell is very bright
		return "00000000" # Black
	else: # Cell is very dark
		return "FFFFFFFF" # White
      
def rgb_to_hex(color):
	# Note: Have color[3] for alpha for future expansion.
	return '%02x%02x%02x' % (color[0], color[1], color[2])

def get_cell_name(x, y):
    if x < 0:
        return f"A{get_row(y)}"  # Left side row numbers
    else:
        return f"{get_column(x + 1)}{get_row(y)}"  # Normal cells and right side row numbers

def get_column(num):
	def divmod_excel(n):
		a, b = divmod(n, 26)
		if b == 0:
			return a - 1, b + 26
		return a, b

	chars = []
	while num > 0:
		num, d = divmod_excel(num)
		chars.append(ascii_uppercase[d - 1])
	return ''.join(reversed(chars)).upper()

def get_row(y):
	return str(y + 1)

def get_used_color_palette(colors, color_map):
	used_colors = []
	used_map = []

	## Get list of used colors
	for x in range(0, len(colors)):
		for y in range(0, len(colors[x])):
			color = colors[x][y]
			color_map_value = color_map[x][y]
			if color not in used_colors:
				used_colors.append(color)
				used_map.append(color_map_value)

	return used_colors, used_map

def check_output_directory():
	if not path.isdir(csv_output_directory):
		mkdir(csv_output_directory)

# DESC: Save the workbook as an excel file
# IN: workbook, file path
# OUT: boolean indicating success
def save_wb(wb, output_file_path):
    try:
        wb.save(output_file_path)
        return True
    except Exception as e:
        print(f"Error saving workbook: {str(e)}")
        return False
    finally:
        try:
            wb.close()
        except Exception as e:
            print(f"Warning: Error during workbook cleanup: {str(e)}")

def cleanup_workbook(wb):
    """Clean up workbook resources"""
    try:
        if wb:
            wb.close()
    except Exception as e:
        print(f"Warning: Error during workbook cleanup: {str(e)}")

# Add cleanup handler for application exit
def on_app_exit():
    """Handle cleanup when application exits"""
    try:
        # Add any additional cleanup needed here
        pass
    except Exception as e:
        print(f"Warning: Error during application cleanup: {str(e)}")

def get_file_name_from_path(file_path):
	return file_path.split("/")[-1]

def export_image_as_excel_pattern(csv_output_directory, include_pixel_numbers = False, include_row_numbers = True):
    column_size = 2.8 # this number is about 20 pixels, same as the default height
    cell_fill_type = 'solid'
    legend_buffer = 2
    output_file_name = "output"
    wb = None

    try:
        image_to_export = image_lvl3

        if image_to_export == None: return

        # get color grid and map from final image
        colors, color_map = get_colors(image_to_export)

        # create worksheet
        wb = Workbook()
        ws = wb.create_sheet(output_file_name, index=0)

        print("Exporting pattern to Excel")

        # if including row numbers column, fill column to left of image with row numbers
        if include_row_numbers:
            #ws.column_dimensions['A'].width = column_size
            for y in range(0, len(colors[0])):
                cell_name = get_cell_name(0, y)  # 0 for leftmost column
                ws[cell_name].value = str(len(colors[0]) - y)  # Fill with descending numbers
                ws[cell_name].alignment = styles.Alignment(horizontal='right')

        # fill worksheet with image
        for x in range(0, len(colors)):
            print("Processing row: " +  str(x) + "/" + str(len(colors)))
            for y in range(0, len(colors[x])):
                cell_color = rgb_to_hex(colors[x][y])
                font_color = get_font_color(colors[x][y])
                cell_symbol = color_map[x][y]
                cell_alignment = styles.Alignment(horizontal='center')
                cell_fill = styles.PatternFill(fill_type=cell_fill_type, start_color=cell_color, end_color=cell_color)
                cell_border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
                cell_font = styles.Font(name='Calibri', bold=False, italic=False, color=font_color)
                # Adjust x coordinate if we have row numbers
                adjusted_x = x + 1 if include_row_numbers else x
                cell_name = get_cell_name(adjusted_x, y)
                ws[cell_name].alignment = cell_alignment
                if include_pixel_numbers: ws[cell_name].value = cell_symbol
                ws[cell_name].fill = cell_fill
                ws[cell_name].border = cell_border
                ws[cell_name].font = cell_font
            # Set column width for the current column
            col_letter = get_column(x + (2 if include_row_numbers else 1))
            ws.column_dimensions[col_letter].width = column_size

        # if including row numbers column, fill column to right of image with row numbers
        if include_row_numbers:
            col_num = len(colors) + 1
            col_letter = get_column(col_num + 1)
            #ws.column_dimensions[col_letter].width = column_size
            for y in range(len(colors[0]) - 1, -1, -1):
                cell_name = get_cell_name(col_num, y)
                ws[cell_name].value = str(len(colors[0]) - y)  # Fill with descending numbers
                ws[cell_name].alignment = styles.Alignment(horizontal='left')

        # add legend (adjust the position to account for the row numbers column)
        used_colors, used_map = get_used_color_palette(colors, color_map)
        width = len(colors)
        legend_start = width + (2 if include_row_numbers else 1)  # Adjust for row numbers column
        for c in range(-1, len(used_colors)):
            if(c == -1):
                ws[get_cell_name(legend_start + legend_buffer, 0)].value = "Color"
                ws[get_cell_name(legend_start + legend_buffer + 1, 0)].value = "HEX"
                ws[get_cell_name(legend_start + legend_buffer + 2, 0)].value = "Red Value"
                ws[get_cell_name(legend_start + legend_buffer + 3, 0)].value = "Green Value"
                ws[get_cell_name(legend_start + legend_buffer + 4, 0)].value = "Blue Value"
                continue        
            color_rgb = used_colors[c]
            color_symbol = used_map[c]
            color_hex = rgb_to_hex(color_rgb)
            font_color = get_font_color(color_rgb)
            cell_font = styles.Font(color=font_color)
            ws[get_cell_name(legend_start + legend_buffer, c + 1)].fill = styles.PatternFill(fill_type=cell_fill_type, start_color=color_hex, end_color=color_hex)
            if include_pixel_numbers: ws[get_cell_name(legend_start + legend_buffer, c + 1)].value = str(color_symbol)
            ws[get_cell_name(legend_start + legend_buffer, c + 1)].font = cell_font
            ws[get_cell_name(legend_start + legend_buffer + 1, c + 1)].value = str(color_hex)
            ws[get_cell_name(legend_start + legend_buffer + 2, c + 1)].value = str(color_rgb[0])
            ws[get_cell_name(legend_start + legend_buffer + 3, c + 1)].value = str(color_rgb[1])
            ws[get_cell_name(legend_start + legend_buffer + 4, c + 1)].value = str(color_rgb[2])

        # save the file
        check_output_directory()
        output_directory = csv_output_directory
        output_file_path = output_directory + "\\" + output_file_name + ".xlsx"
        save_success = save_wb(wb, output_file_path)
        if save_success:
            print("Export complete!")
            print(f"File '{output_file_name}.xlsx' created at location: '{output_file_path}'")
        else:
            print(output_file_name + " save failed")
            messagebox.showinfo(error_box_header, "Error: Save failed. Make sure file '" + get_file_name_from_path(output_file_name) + "' is not already open on computer.")
    except Exception as e:
        print(f"Error during export: {str(e)}")
        messagebox.showinfo(error_box_header, f"Error during export: {str(e)}")
    finally:
        if wb:
            cleanup_workbook(wb)
            


# Import a pattern from Excel, specifying the rectangular region by start_cell and end_cell (e.g., "B1", "BX75").
def import_pattern_from_excel(start_cell, end_cell):
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    # Prompt user to select Excel file
    filepath = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not filepath:
        print("No file selected.")
        return None

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        # Parse start and end cell coordinates
        start_col_name, start_row = coordinate_from_string(start_cell.upper())
        end_col_name, end_row = coordinate_from_string(end_cell.upper())
        start_col = column_index_from_string(start_col_name)
        end_col = column_index_from_string(end_col_name)
        start_row = int(start_row)
        end_row = int(end_row)

        # Ensure correct order (top-left to bottom-right)
        min_row, max_row = min(start_row, end_row), max(start_row, end_row)
        min_col, max_col = min(start_col, end_col), max(start_col, end_col)

        width = max_col - min_col + 1
        height = max_row - min_row + 1

        # Read cell colors and build pixel data
        pixels = []
        for row in range(min_row, max_row + 1):
            row_pixels = []
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                # Try to get the cell's fill color
                fill = cell.fill
                rgb = (255, 255, 255)  # default to white
                if fill and fill.fill_type is not None and fill.start_color is not None:
                    color = fill.start_color
                    # openpyxl color can be indexed or rgb
                    if color.type == "rgb" and color.rgb is not None:
                        hexstr = color.rgb
                        # openpyxl may return ARGB, so skip first two chars if length 8
                        if len(hexstr) == 8:
                            hexstr = hexstr[2:]
                        try:
                            rgb = tuple(int(hexstr[i:i+2], 16) for i in (0, 2, 4))
                        except Exception:
                            rgb = (255, 255, 255)
                row_pixels.append(rgb)
            pixels.append(row_pixels)

        # Create image from pixel data
        img = Image.new("RGB", (width, height))
        for y in range(height):
            for x in range(width):
                img.putpixel((x, y), pixels[y][x])

        print("Pattern imported from Excel successfully!")
        print(f"File path: '{filepath}'")

        # test import by replacing image_lvl0
        # image_lvl0 = img
        # update_image_display(image_lvl0, image_lvl0_image_label)

        return img

    except Exception as e:
        print(f"Error importing pattern from Excel: {e}")
        return None





# ---------- GUI Setup ----------

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.geometry(f"{window_width}x{window_height}")
app.title(window_title)
# configure to use grid
app.grid_rowconfigure(0, weight=1) # make row 0 expandable
app.grid_columnconfigure(0, weight=1) # make column 0 expandable

# ---------- UI Frame Structure ----------

frame_app = ctk.CTkFrame(app)
frame_app.grid(row=0, column=0, padx=5, pady=5, sticky="nsew") # new = stick to sides: north, south, east, west
frame_app.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_app.grid_columnconfigure(0, weight=1) # set column 0 to expandable
# set columns to be expandable so they appear evenly spaced across the window
frame_app.grid_columnconfigure(0, weight=1)
frame_app.grid_columnconfigure(1, weight=1)
frame_app.grid_columnconfigure(2, weight=1)

# frame for column 0
frame_col0 = ctk.CTkFrame(frame_app, fg_color="transparent")
frame_col0.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
frame_col0.grid_columnconfigure(0, weight=1) # set column 0 to expandable
frame_col0.grid_rowconfigure(0, weight=1) # set row 0 to expandable

frame_col1 = ctk.CTkFrame(frame_app, fg_color="transparent")
frame_col1.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
frame_col1.grid_columnconfigure(0, weight=1) # set column 0 to expandable
frame_col1.grid_rowconfigure(0, weight=1) # set row 0 to expandable

frame_col2 = ctk.CTkFrame(frame_app, fg_color="transparent")
frame_col2.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")
frame_col2.grid_columnconfigure(0, weight=1) # set column 0 to expandable
frame_col2.grid_rowconfigure(0, weight=1) # set row 0 to expandable

# ---------- Frame: col0 ----------

# image editting frame for column 0, for image lvl0 display and all associated controls
frame_col0_image_editing = ctk.CTkFrame(frame_col0, fg_color=("gray75", "gray25"))
frame_col0_image_editing.grid(row=0, column=0, padx=5, pady=5, sticky="new")
frame_col0_image_editing.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# text label for step 1
label_step1 = ctk.CTkLabel(frame_col0_image_editing, text="Step 1", font=("Arial", 24))
label_step1.grid(row=0, column=0, padx=5, pady=5)

# image_lvl0 (original) frame for image
frame_image_lvl0 = ctk.CTkFrame(frame_col0_image_editing, width=500, height=500)
frame_image_lvl0.grid(row=1, column=0, padx=5, pady=5)
frame_image_lvl0.grid_propagate(False)
# make needed columns and rows expandable so things can be centered
frame_image_lvl0.grid_columnconfigure(0, weight=1)
frame_image_lvl0.grid_rowconfigure(0, weight=1)

# image_lvl0 (original) image label
image_lvl0_image_label = ctk.CTkLabel(frame_image_lvl0, text="")
image_lvl0_image_label.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# button to select an image file
file_button = ctk.CTkButton(frame_col0_image_editing, text="Select Image File", command = lambda: select_file())
file_button.grid(row=2, column=0, padx=5, pady=5)

# console frame
frame_console = ctk.CTkFrame(frame_col0, fg_color=("gray75", "gray25"))
frame_console.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
frame_console.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_console.grid_columnconfigure(0, weight=1) # set column 0 to expandable
# console text label
console_text_label = ctk.CTkLabel(frame_console, text="Console")
console_text_label.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
# console text box
console_text_box = ctk.CTkTextbox(frame_console, state="disabled")
console_text_box.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

# redirect stdout to the console textbox
sys.stdout = TextBoxRedirector(console_text_box)

# ---------- Frame: col1 ----------

# image editting frame for column 1, for image lvl1 display and all associated controls
frame_col1_image_editing = ctk.CTkFrame(frame_col1, fg_color=("gray75", "gray25"))
frame_col1_image_editing.grid(row=0, column=0, padx=5, pady=5, sticky="new")
frame_col1_image_editing.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# text label for step 2
label_step1 = ctk.CTkLabel(frame_col1_image_editing, text="Step 2", font=("Arial", 24))
label_step1.grid(row=0, column=0, padx=5, pady=5)

# image_lvl1 frame for image
frame_image_lvl1 = ctk.CTkFrame(frame_col1_image_editing, width=500, height=500)
frame_image_lvl1.grid(row=1, column=0, padx=5, pady=5)
frame_image_lvl1.grid_propagate(False)
# make needed columns and rows expandable so things can be centered
frame_image_lvl1.grid_columnconfigure(0, weight=1)
frame_image_lvl1.grid_rowconfigure(0, weight=1)

# image_lvl1 image label
image_lvl1_image_label = ctk.CTkLabel(frame_image_lvl1, text="")
image_lvl1_image_label.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# frame for value sliders
frame_image_value_sliders = ctk.CTkFrame(frame_col1_image_editing)
frame_image_value_sliders.grid(row=2, column=0, padx=5, pady=5)

# dummy content 1
# frame_dummy_preprocess1 = ctk.CTkFrame(frame_col1_image_editing)
# frame_dummy_preprocess1.grid(row=3, column=0, padx=5, pady=5)
# label_dummy_preprocess1 = ctk.CTkLabel(frame_dummy_preprocess1, text="put more controls here")
# label_dummy_preprocess1.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# ---------- Frame: col2 ----------

# image editting frame for column 2, for image lvl2 display and all associated controls
frame_col2_image_editing = ctk.CTkFrame(frame_col2, fg_color=("gray75", "gray25"))
frame_col2_image_editing.grid(row=0, column=0, padx=5, pady=5, sticky="new")
frame_col2_image_editing.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# text label for step 3
label_step2 = ctk.CTkLabel(frame_col2_image_editing, text="Step 3", font=("Arial", 24))
label_step2.grid(row=0, column=0, padx=5, pady=5)

# image_lvl2 frame for image
frame_image_lvl2 = ctk.CTkFrame(frame_col2_image_editing, width=500, height=500)
frame_image_lvl2.grid(row=1, column=0, padx=5, pady=5)
frame_image_lvl2.grid_propagate(False)
# make needed columns and rows expandable so things can be centered
frame_image_lvl2.grid_columnconfigure(0, weight=1)
frame_image_lvl2.grid_rowconfigure(0, weight=1)

# image_lvl2 image label
image_lvl2_image_label = ctk.CTkLabel(frame_image_lvl2, text="")
image_lvl2_image_label.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")


# frame for pixelate controls
frame_pixelate = ctk.CTkFrame(frame_col2_image_editing)
frame_pixelate.grid(row=2, column=0, padx=5, pady=5)

# frame for export to excel controls
frame_export_to_excel = ctk.CTkFrame(frame_col1, fg_color=("gray75", "gray25"))
frame_export_to_excel.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
frame_export_to_excel.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_export_to_excel.grid_rowconfigure(1, weight=1) # set row 1 to expandable
frame_export_to_excel.grid_columnconfigure(0, weight=1) # set column 0 to expandable
frame_export_to_excel.grid_columnconfigure(1, weight=1) # set column 1 to expandable

# frame for pixel shift controls
frame_pixel_shift = ctk.CTkFrame(frame_col2, fg_color=("gray75", "gray25"))
frame_pixel_shift.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
frame_pixel_shift.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_pixel_shift.grid_rowconfigure(1, weight=1) # set row 1 to expandable
frame_pixel_shift.grid_columnconfigure(0, weight=1) # set column 0 to expandable
frame_pixel_shift.grid_columnconfigure(1, weight=1) # set column 1 to expandable

# ---------- Frame: Value Sliders ----------

# reset sliders button
reset_button = ctk.CTkButton(frame_image_value_sliders, text="Reset Sliders", command = lambda: reset_sliders())
reset_button.grid(row=0, column=0, padx=5, pady=5)

# sliders: brightness, contrast, saturation
frame_slider = ctk.CTkFrame(frame_image_value_sliders, fg_color="transparent")
frame_slider.grid(row=0, column=1, padx=5, pady=5)

slider_defaults = {"brightness": 1.0, "contrast": 1.0, "saturation": 1.0}

def create_slider(label_text, var_name, row):
    label = ctk.CTkLabel(frame_slider, text=label_text)
    label.grid(row=row, column=0, sticky="w")
    slider = ctk.CTkSlider(
        frame_slider,
        from_=0.2,
        to=2.0,
        number_of_steps=100,
        command = lambda v: update_all_levels(),
    )
    slider.set(slider_defaults[var_name])
    slider.grid(row=row, column=1, pady=5, padx=5)
    return slider

brightness_slider = create_slider("Brightness", "brightness", 0)
contrast_slider   = create_slider("Contrast", "contrast", 1)
saturation_slider = create_slider("Saturation", "saturation", 2)

# ---------- Frame: Pixelate ----------

# update button
update_button = ctk.CTkButton(frame_pixelate, text="Update", command = lambda: update_all_levels())
update_button.grid(row=1, column=0, padx=5, pady=5)

# entry boxes
frame_entry = ctk.CTkFrame(frame_pixelate)
frame_entry.grid(row=0, column=0, padx=5, pady=5)

def create_entry(label_text, default_text, row):
    lbl = ctk.CTkLabel(frame_entry, text=label_text)
    lbl.grid(row=row, column=0, sticky="w")
    entry = ctk.CTkEntry(frame_entry)
    entry.insert(0, default_text)
    entry.grid(row=row, column=1, pady=5, padx=5)
    entry.bind('<Return>', lambda e: update_all_levels()) # update images when user presses Enter
    entry.bind('<FocusOut>', lambda e: update_all_levels()) # update images when user leaves text box
    return entry

width_entry  = create_entry("Width (# of stitches)", "75", 0)
height_entry = create_entry("Height (# of stitches)", "75", 1)
colors_entry = create_entry("Number of Colors", "3", 2)

# ---------- Frame: Export to Excel ----------

# label for title
label_step4 = ctk.CTkLabel(frame_export_to_excel, text="Step 4", font=("Arial", 24))
label_step4.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# frame for export button
frame_export_button = ctk.CTkFrame(frame_export_to_excel, fg_color="transparent")
frame_export_button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
frame_export_button.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_export_button.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# frame for controls
frame_export_controls = ctk.CTkFrame(frame_export_to_excel, fg_color="transparent")
frame_export_controls.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
frame_export_controls.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_export_controls.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# export to excel button
export_image_as_pattern_button = ctk.CTkButton(frame_export_button, text="Export Pattern to Excel", command = lambda: export_image_as_excel_pattern(csv_output_directory, include_pixel_numbers = include_cells_var.get(), include_row_numbers = include_rownums_var.get()))
export_image_as_pattern_button.grid(row=0, column=0, padx=5, pady=5)

# checkbox: include cell numbers
include_cells_var = ctk.BooleanVar()
checkbox_color_nums = ctk.CTkCheckBox(frame_export_controls, text="Include color numbers in cells", variable=include_cells_var)
checkbox_color_nums.grid(row=0, column=0, padx=5, pady=5, sticky="nsw")

# checkbox: include row numbers
include_rownums_var = ctk.BooleanVar(value=True)
checkbox_include_rownums = ctk.CTkCheckBox(frame_export_controls, text="Include row numbers", variable=include_rownums_var)
checkbox_include_rownums.grid(row=1, column=0, padx=5, pady=5, sticky="nsw")

# ---------- Frame: Pixel Shift ----------

# label for title
label_step5 = ctk.CTkLabel(frame_pixel_shift, text="Step 5", font=("Arial", 24))
label_step5.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# frame for load pattern button
frame_load_pattern_button = ctk.CTkFrame(frame_pixel_shift, fg_color="transparent")
frame_load_pattern_button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
frame_load_pattern_button.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_load_pattern_button.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# frame for controls
frame_controls = ctk.CTkFrame(frame_pixel_shift, fg_color="transparent")
frame_controls.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
frame_controls.grid_rowconfigure(0, weight=1) # set row 0 to expandable
frame_controls.grid_columnconfigure(0, weight=1) # set column 0 to expandable

# export to excel button
load_pattern_button = ctk.CTkButton(frame_load_pattern_button, text="Load Pattern from Excel", command = lambda: import_pattern_from_excel('B1', 'BX75'))
load_pattern_button.grid(row=0, column=0, padx=5, pady=5)

# checkbox: some checkbox
include_cells_var = ctk.BooleanVar()
checkbox_dummy = ctk.CTkCheckBox(frame_controls, text="some checkbox", variable=include_cells_var)
checkbox_dummy.grid(row=0, column=0, padx=5, pady=5, sticky="nsw")

# checkbox: some checkbox
include_rownums_var = ctk.BooleanVar(value=True)
checkbox_dummy1 = ctk.CTkCheckBox(frame_controls, text="some checkbox2", variable=include_rownums_var)
checkbox_dummy1.grid(row=1, column=0, padx=5, pady=5, sticky="nsw")

# ---------- Launch ----------
app.mainloop()

