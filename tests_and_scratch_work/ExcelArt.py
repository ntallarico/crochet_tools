'''
Excel Image Art Generator
 
Converts images to Excel color grids with configurable resolution using pixel averaging.
'''

import openpyxl as xls
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
from tkinter import Tk, Label


DEFAULT_STEP = 5 # higher the step, lower the resolution. unsure how this works yet
CELL_WIDTH = 3
ROW_HEIGHT = 20


def process(path: str, step: int = DEFAULT_STEP) -> list[list[tuple[int, int, int]]]:
    '''
    Process image into color blocks

    Args:
        path: Path to source image
        step: Pixel grouping size (controls output resolution)

    Returns:
        2D list of RGB color tuples
    '''

    image = Image.open(path)
    w, h = image.size
    colors = []

    for y in range(0, h, step):
        row_colors = []
        for x in range(0, w, step):
            # Calculate block boundaries
            box = (
                x,
                y,
                min(x+step, w),
                min(y+step, h)
            )

            block = image.crop(box)

            row_colors.append(calculate_average_color(block))
        colors.append(row_colors)

    return colors


def calculate_average_color(block: Image.Image) -> tuple[int, int, int]:
    '''Calculate average color of an image block'''
    r_total = g_total = b_total = 0
    pixel_count = block.width * block.height
    for x in range(block.width):
        for y in range(block.height):
            r, g, b = block.getpixel((x, y))[:3]
            r_total += r
            g_total += g
            b_total += b
    return (r_total // pixel_count, g_total // pixel_count, b_total // pixel_count)


def create_excel(colors: list[list[tuple[int, int, int]]], result_path: str):
    """
    Generate Excel file from color grid

    Args:
        colors: 2D list of RGB tuples from process()
        output_path: Path to save Excel file
    """

    wb = xls.Workbook()
    ws = wb.active


    for i, row in enumerate(colors, 1):
        row_hex_color_list = []
        for j, color in enumerate(row, 1):

            ws.column_dimensions[get_column_letter(j)].width = CELL_WIDTH
            ws.row_dimensions[i].height = ROW_HEIGHT

            hex_color = f"{color[0]:02X}{color[1]:02X}{color[2]:02X}"

            ws.cell(row=i, column=j).fill = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type="solid"
            )

    wb.save(result_path)


def tranlate_color_array_rgb_to_hex(colors: list[list[tuple[int, int, int]]]):
    image_hex_color_array = []

    for i, row in enumerate(colors, 1):
        row_hex_color_list = []
        for j, color in enumerate(row, 1):

            hex_color = f"{color[0]:02X}{color[1]:02X}{color[2]:02X}"
            row_hex_color_list.append(hex_color)

        image_hex_color_array.append(row_hex_color_list)

    return image_hex_color_array


def enlarge_image(rgb_data, scale=5):
    enlarged_data = []

    for row in rgb_data:
        # Create a new row for the enlarged image
        enlarged_row = []
        for rgb in row:
            # Add the RGB value to the enlarged row 5 times (for 5 pixels)
            enlarged_row.extend([rgb] * scale)
        
        # Add the enlarged row 5 times to the enlarged data (for 5 pixels in height)
        for _ in range(scale):
            enlarged_data.append(enlarged_row)

    return enlarged_data


def display_image(rgb_data):
    # Create an image from the RGB data
    height = len(rgb_data)
    width = len(rgb_data[0])
    image = Image.new("RGB", (width, height))

    for y in range(height):
        for x in range(width):
            image.putpixel((x, y), rgb_data[y][x])

    # Create a Tkinter window
    root = Tk()
    root.title("Image Display")

    # Convert the image to a format Tkinter can use
    tk_image = ImageTk.PhotoImage(image)

    # Create a label to display the image
    label = Label(root, image=tk_image)
    label.pack()

    # Start the Tkinter event loop
    root.mainloop()



def main():
    color_array_rgb = process(r"C:\Users\Nick\Stuff\Projects\Crochet Tools\crochet_tools\input_output\img.jpg", step=10)

    create_excel(color_array_rgb, r"C:\Users\Nick\Stuff\Projects\Crochet Tools\crochet_tools\input_output\output.xlsx")

    for row in color_array_rgb: print(row)    

    color_array_rgb_enlarged = enlarge_image(color_array_rgb, 10)

    #display_image(color_array_rgb)
    display_image(color_array_rgb_enlarged)

    # color_array_hex = tranlate_color_array_rgb_to_hex(color_array_rgb)
    # for row in color_array_hex: print(row)

    
    

    

if __name__ == "__main__":
    main()
    
